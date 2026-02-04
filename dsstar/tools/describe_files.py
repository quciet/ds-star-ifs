from __future__ import annotations

import csv
import importlib.util
import json
from pathlib import Path
from typing import Any, Dict, List


def _infer_type(value: str) -> str:
    if value == "":
        return "empty"
    try:
        int(value)
        return "int"
    except ValueError:
        pass
    try:
        float(value)
        return "float"
    except ValueError:
        pass
    return "str"


def _describe_csv(path: Path) -> Dict[str, Any]:
    rows: List[List[str]] = []
    with path.open(newline="", encoding="utf-8", errors="replace") as handle:
        reader = csv.reader(handle)
        for _ in range(6):
            try:
                rows.append(next(reader))
            except StopIteration:
                break
    header = rows[0] if rows else []
    sample_rows = rows[1:6] if rows else []
    type_hints = []
    if sample_rows and header:
        for col_idx, name in enumerate(header):
            sample_values = [row[col_idx] for row in sample_rows if col_idx < len(row)]
            inferred = {_infer_type(value) for value in sample_values if value is not None}
            type_hints.append({"column": name, "types": sorted(inferred)})
    return {"type": "csv", "header": header, "sample_rows": sample_rows, "type_hints": type_hints}


def _describe_text(path: Path) -> Dict[str, Any]:
    content = path.read_text(encoding="utf-8", errors="replace")
    snippet = content[:2000]
    return {"type": "text", "length": len(content), "snippet": snippet}


def _describe_json(path: Path) -> Dict[str, Any]:
    content = path.read_text(encoding="utf-8", errors="replace")
    snippet = content[:2000]
    try:
        parsed = json.loads(content)
        summary = type(parsed).__name__
    except json.JSONDecodeError:
        summary = "invalid_json"
    return {"type": "json", "summary": summary, "length": len(content), "snippet": snippet}


def _describe_xlsx(path: Path) -> Dict[str, Any]:
    if importlib.util.find_spec("openpyxl") is None:
        return {
            "type": "xlsx",
            "warning": "openpyxl not installed; cannot read sheets.",
        }
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    sheets_info = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append([str(cell) if cell is not None else "" for cell in row])
            if idx >= 5:
                break
        sheets_info[sheet] = {"sample_rows": rows}
    return {"type": "xlsx", "sheets": sheets_info}


def describe_files(paths: List[str], output_path: Path) -> Dict[str, Any]:
    descriptions: Dict[str, Any] = {"files": {}, "warnings": []}
    for raw in paths:
        path = Path(raw)
        if not path.exists():
            descriptions["warnings"].append(f"Missing file: {raw}")
            continue
        suffix = path.suffix.lower()
        try:
            if suffix == ".csv":
                info = _describe_csv(path)
            elif suffix == ".xlsx":
                info = _describe_xlsx(path)
            elif suffix in {".json"}:
                info = _describe_json(path)
            elif suffix in {".txt", ".md"}:
                info = _describe_text(path)
            else:
                info = _describe_text(path)
            descriptions["files"][raw] = info
        except Exception as exc:  # pylint: disable=broad-except
            descriptions["warnings"].append(f"Failed to describe {raw}: {exc}")
    output_path.write_text(json.dumps(descriptions, indent=2), encoding="utf-8")
    return descriptions
