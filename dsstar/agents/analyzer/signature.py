from __future__ import annotations

import csv
import sqlite3
from pathlib import Path
from typing import Dict, List


def _size_bucket(size: int) -> str:
    if size < 10_000:
        return "xs"
    if size < 100_000:
        return "s"
    if size < 1_000_000:
        return "m"
    if size < 10_000_000:
        return "l"
    return "xl"


def _first_non_empty_lines(path: Path, limit: int = 5) -> List[str]:
    lines: List[str] = []
    try:
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for raw in handle:
                line = raw.strip()
                if not line:
                    continue
                lines.append(line)
                if len(lines) >= limit:
                    break
    except Exception:
        return []
    return lines


def _guess_delimiter(lines: List[str]) -> str:
    candidates = [",", "\t", ";", "|"]
    scores = {c: 0 for c in candidates}
    for line in lines:
        for c in candidates:
            scores[c] += line.count(c)
    return max(scores, key=scores.get) if lines else "none"


def _excel_probe(path: Path) -> Dict[str, str]:
    meta: Dict[str, str] = {"sheet_count": "0", "first_sheet": "none", "first_row_cells": "0"}
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        meta["sheet_count"] = str(len(names))
        if names:
            meta["first_sheet"] = names[0]
            ws = wb[names[0]]
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            meta["first_row_cells"] = str(len([v for v in first_row if v is not None]))
    except Exception:
        pass
    return meta


def _sqlite_probe(path: Path) -> Dict[str, str]:
    meta: Dict[str, str] = {"tables": "none"}
    try:
        conn = sqlite3.connect(str(path))
        cur = conn.cursor()
        rows = cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name LIMIT 5").fetchall()
        conn.close()
        names = [str(r[0]) for r in rows]
        meta["tables"] = ",".join(names) if names else "none"
    except Exception:
        pass
    return meta


def compute_signature(path: Path) -> str:
    suffix = path.suffix.lower() or "none"
    size_bucket = _size_bucket(path.stat().st_size)
    parts = [f"ext={suffix}", f"size={size_bucket}"]

    if suffix in {".csv", ".tsv", ".txt", ".json"}:
        lines = _first_non_empty_lines(path)
        delim = _guess_delimiter(lines)
        header = lines[0][:120] if lines else "none"
        col_count = "0"
        if lines:
            try:
                col_count = str(len(next(csv.reader([lines[0]], delimiter=delim if delim != "none" else ","))))
            except Exception:
                col_count = "0"
        parts.extend([f"delim={delim}", f"cols={col_count}", f"header={header}"])
    elif suffix in {".xlsx", ".xls", ".xlsm"}:
        meta = _excel_probe(path)
        parts.extend([f"sheet_count={meta['sheet_count']}", f"first_sheet={meta['first_sheet']}", f"first_row_cells={meta['first_row_cells']}"])
    elif suffix in {".db", ".sqlite", ".sqlite3"}:
        meta = _sqlite_probe(path)
        parts.append(f"tables={meta['tables']}")

    return "|".join(parts)


def probe_sample(path: Path, max_lines: int = 50) -> str:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt", ".json"}:
        try:
            lines: List[str] = []
            with path.open("r", encoding="utf-8", errors="replace") as handle:
                for idx, line in enumerate(handle):
                    if idx >= max_lines:
                        break
                    lines.append(line.rstrip("\n"))
            return "\n".join(lines)
        except Exception as exc:
            return f"text_probe_failed: {exc}"
    if suffix in {".xlsx", ".xls", ".xlsm"}:
        return str(_excel_probe(path))
    if suffix in {".db", ".sqlite", ".sqlite3"}:
        return str(_sqlite_probe(path))
    return f"file_size={path.stat().st_size}"
