from __future__ import annotations

import csv
import json
import sqlite3
from pathlib import Path


def _safe_text(path: Path, max_chars: int = 800) -> str:
    return path.read_text(encoding="utf-8", errors="replace")[:max_chars].replace("\n", " ")


def describe_file(path: str) -> str:
    p = Path(path)
    try:
        ext = p.suffix.lower()
        lines = [f"FILE={p}", f"EXT={ext}", f"SIZE={p.stat().st_size}"]

        if ext in {".csv", ".tsv"}:
            delim = "\t" if ext == ".tsv" else ","
            with p.open("r", encoding="utf-8", errors="replace") as handle:
                reader = csv.reader(handle, delimiter=delim)
                rows = list(reader)
            if rows:
                lines.append("COLUMNS=" + ",".join([str(c) for c in rows[0]]))
                lines.append("ROW_COUNT=" + str(max(len(rows) - 1, 0)))
                if len(rows) > 1:
                    lines.append("FIRST_DATA_ROW=" + ",".join([str(c) for c in rows[1]]))
        elif ext in {".json"}:
            payload = json.loads(p.read_text(encoding="utf-8", errors="replace"))
            lines.append("JSON_TYPE=" + type(payload).__name__)
            if isinstance(payload, dict):
                lines.append("TOP_KEYS=" + ",".join(sorted(payload.keys())[:20]))
            elif isinstance(payload, list):
                lines.append("LIST_LEN=" + str(len(payload)))
        elif ext in {".db", ".sqlite", ".sqlite3"}:
            conn = sqlite3.connect(str(p))
            cur = conn.cursor()
            tables = [r[0] for r in cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()]
            lines.append("TABLES=" + ",".join(tables))
            for t in tables[:3]:
                cols = [r[1] for r in cur.execute(f"PRAGMA table_info('{t}')").fetchall()]
                lines.append(f"TABLE_{t}_COLUMNS=" + ",".join(cols))
            conn.close()
        elif ext in {".xlsx", ".xls", ".xlsm"}:
            try:
                from openpyxl import load_workbook  # type: ignore

                wb = load_workbook(p, read_only=True, data_only=True)
                sheets = list(wb.sheetnames)
                lines.append("SHEETS=" + ",".join(sheets))
                if sheets:
                    ws = wb[sheets[0]]
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                    lines.append("FIRST_SHEET_HEADER=" + ",".join([str(v) for v in first_row if v is not None]))
            except Exception as exc:  # pylint: disable=broad-except
                lines.append(f"XLSX_PARSE_WARNING={exc}")
                lines.append("SNIPPET=" + _safe_text(p, 200))
        else:
            lines.append("SNIPPET=" + _safe_text(p))

        return "\n".join(lines)
    except Exception as exc:  # pylint: disable=broad-except
        return f"FAILED TO DESCRIBE: {exc}"
